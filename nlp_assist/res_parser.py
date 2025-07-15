# -*- coding: utf-8 -*-
"""
Created on 2020/1/21

@author: fk506cni=unkodaisuki!

result parse class.
"""
from tqdm import tqdm
import pandas as pd
import numpy as np
import itertools as it
import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import nsp_settings

import openpyxl  as xl

class res_parser():

    def __init__(self, x):
        print("parser activated!")
        self.x = x
        self.rc = res_checker(x=x)
        self.dt1 = x.dt1
        self.dt2 = x.dt2
        self.dt3 = x.dt3

    def saveRes_asCSV(self, res, path):
        print("mada")
        n = []
        v = []
        for k, i in tqdm(res.items()):
            n.append(k)
            v.append(i)
        df = pd.DataFrame(columns=["node", "val"])
        df["node"] = n
        df["val"] = v
        df.to_csv(path, sep=",", index=None)

    def readRes_fromCSV(self, path):
        res_df = pd.read_csv(path, sep=",")
        self.res_df = res_df.copy()
        res_dict = dict()
        for r in tqdm(res_df.itertuples()):
            res_dict[r.node] = r.val
        self.res = res_dict

    def setRes(self, res):
        self.res = res
        res_df = pd.DataFrame(columns=["node", "val"])
        n = []
        v = []
        for k, i in tqdm(res.items()):
            n.append(k)
            v.append(n)
        res_df["node"] = n
        res_df["val"] = v
        self.res_df = res_df.copy()

    def prepRes(self, res, hideProgressBar=False):
        print("checking...")

        pos_dt1 = []
        pos_dt2 = []
        pos_dt3 = []
        l_0 = []
        l_1 = []
        l_2 = []
        l_3 = []
        l_4 = []
        l_5 = []
        vals = []

        for k, i in tqdm(res.items(), disable=hideProgressBar):
            if ("*" not in k) and ("dd" not in k) and ("dn" not in k) and ("dv" not in k):

                s1, s2, s3, s4, s5 = k.split("_")
                l_0.append(k)
                l_1.append(s1)
                l_2.append(s2)
                l_3.append(s3)
                l_4.append(s4)
                l_5.append(s5)
                vals.append(i)
            else:
                pass

        df = pd.DataFrame(
            columns=["l_0", "l_1", "l_2", "l_3", "l_4", "l_5", "vals"])
        df["l_0"] = l_0
        df["l_1"] = l_1
        df["l_2"] = l_2
        df["l_3"] = l_3
        df["l_4"] = l_4
        df["l_5"] = l_5
        df["vals"] = vals

        df["l_012"] = df["l_1"] + "_" + df["l_2"] + "_0_0_0"
        df["l_0123"] = df["l_1"] + "_" + df["l_2"] + "_" + df["l_3"] + "_0_0"

        # df1 = df.query("l_3 == '0'")
        df1 = df.query("vals==1").loc[:, ["vals", "l_012"]].drop_duplicates()
        # df2 = df.query("l_4 == '0'")
        df2 = df.query("vals==1").loc[:, ["vals", "l_0123"]].drop_duplicates()
        df3 = df.query("l_4 != '0'")

        self.dt1_mg = pd.merge(
            self.dt1, df1, left_on="var_name", right_on="l_012", how="left")
        self.dt1_mg["vals"] = self.dt1_mg["vals"].fillna(0)
        self.dt2_mg = pd.merge(
            self.dt2, df2, left_on="var_name", right_on="l_0123", how="left")
        self.dt2_mg["vals"] = self.dt2_mg["vals"].fillna(0)
        self.dt3_mg = pd.merge(
            self.dt3, df3, left_on="var_name", right_on="l_0")

    def save3cond_asCSV(self, prefix="mg_cond_"):
        self.dt1_mg.to_csv(prefix+"dt1.csv", sep=",",
                           index=None, encoding='utf_8_sig')
        self.dt2_mg.to_csv(prefix + "dt2.csv", sep=",",
                           index=None, encoding='utf_8_sig')
        self.dt3_mg.to_csv(prefix + "dt3.csv", sep=",",
                           index=None, encoding='utf_8_sig')


class res_checker():

    setting = nsp_settings.nsp_settings()
    echo_bool = True
    clps_dict_org = {
        "date_number": -1,
        "members": "",
        "dutys": "",
        "shifts": "",
        "info": "",
        "class": "",
        "comments": ""
    }
    log_cols = list(clps_dict_org.keys())

    log_df = pd.DataFrame(columns=log_cols)
    #clps_list = []

    log_class = setting.log_class
    check_list = setting.check_list
    # check_list = ["3dts", "1p1", "cnight", "cwork", "fseq", "nseq", "oAb", "wres", "eq", "eqn", "eqg3", "eqg2", "lfbd","lfbd", "lnd", "lcnt", "lcnn"]

    hideProgressBar = True

    def pp(self, *args):
        if self.echo_bool:
            print(*args)

    def __init__(self, x, hideProgressBar=True, echo_bool=False):
        self.x = x
        self.hideProgressBar = hideProgressBar
        self.echo_bool = echo_bool
        self.pp("checker activated!")
        self.clps_list = []

    # リファクタリング後の関数群

    def check_one_by_one_task(self, dt3):
        # rp.dt3_mg.loc[:, ["name_ind", "date_number"]].drop_duplicates()
        bol = False
        df_dup = dt3 \
            .groupby(["name_ind", "date_number", "name"])["vals"]\
            .sum()\
            .reset_index()\
            .query("vals >1")

        if len(df_dup) > 0:
            bol = True
            for r in df_dup.itertuples():
                clps_dict = self.clps_dict_org.copy()
                clps_dict["members"] = str(r.name)
                clps_dict["date_number"] = r.date_number
                clps_dict["class"] = "1by1"
                clps_dict["info"] = str(r.vals)
                self.pp(clps_dict)
                self.clps_list.append(clps_dict)

        # for r in dt3.loc[:, ["name_ind", "date_number", "name"]].drop_duplicates().itertuples():
        #     df_r = dt3.query("name_ind == @r.name_ind and date_number == @r.date_number and vals ==1")
        #     if df_r["vals"].sum() > 1:
        #         self.pp("sum over, not 1 by 1 task")
        #         dup_task = df_r["DutyLabel"].to_list()
        #         status = str(r.date_number) + "_" + str(r.name_ind)+"_"+r.name
        #         clps_dict = self.clps_dict_org.copy()
        #         clps_dict["members"] = str(r.name_ind)
        #         clps_dict["dutys"] = dup_task

        #         self.pp(status)
        #         self.pp(dup_task)
        #         tmp_se = pd.Series(["1by1", status, "not_one_by_one"], index=self.log_cols)
        #         self.log_df = self.log_df.append(tmp_se, ignore_index=True)
        #         bol =True
        if bol:
            self.pp("not 1 by 1")
        else:
            self.pp("is in one by one")

    def check_contWork(self, dt3):
        self.pp("checking continous working...")
        bol = False
        for m in tqdm(self.x.df_MemberCapacity["name"], disable=self.hideProgressBar):

            # self.pp(m)
            # group by してmergeしたほうが絶対早いができていない。

            m_capa = int(self.x.df_MemberCapacity.query(
                "name == @m")["CapableContinuousWorks"])
            max_date = int(self.x.max_date)
            # # self.pp(m_capa)
            dt3m = dt3.query(
                "name == @m and vals ==1").drop_duplicates(subset="date_number")
            #     and date_number < @max_date -@m_capa
            for d in range(0, max_date - m_capa + 1):

                #         self.pp(d)
                # var_tuple = dt1m.query("@d+1 <= date_number <= @d +@m_capa")["var_name"].to_list()
                var_sum = dt3m.query(
                    "@d+1 <= date_number <= @d +@m_capa")["vals"].sum()
                if var_sum > m_capa:

                    bol = True
                    clps_dict = self.clps_dict_org.copy()
                    clps_dict["members"] = str(m)
                    clps_dict["date_number"] = -1
                    clps_dict["class"] = "cwork"

                    status = "continous_work_over: " + m + \
                        ", start:" + str(d+1)+", end:"+str(d + m_capa)
                    clps_dict["info"] = status
                    self.clps_list.append(status)
                    #tmp_se = pd.Series(["cwork", status, "over_work"],index=self.log_cols)
                    #self.log_df = self.log_df.append(tmp_se, ignore_index=True)
                    self.pp(status)
        if bol:
            self.pp("over contineous work")

    # def check_dt1_to_dt2(self, dt1, dt2):
    #     self.pp("check dt1 to dt2")
    #     bol = False
    #     for r in tqdm(dt1.itertuples(), disable=self.hideProgressBar):
    #         name_ind = r.name_ind
    #         name = r.name
    #         date_number = r.date_number
    #         # var_name = r.var_name
    #         var_vals = r.vals
    #         var_varls_dt2 = dt2.query("name_ind == @name_ind & date_number == @date_number")["vals"].sum()
    #         if var_vals != var_varls_dt2:
    #             bol = True
    #             # self.pp("dt1 not eq dt2")
    #             # self.pp(var_vals)
    #             # self.pp(var_varls_dt2)
    #             status = "dt1 not eq dt2: " + name + " " + str(date_number)
    #             tmp_se = pd.Series([self.log_class[2], status, "dt1_not_eq_dt2"],index=self.log_cols)
    #             self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #             self.pp(status)
    #     if bol:
    #         self.pp("dt1 not equal to dt2..l.")
    #
    #
    # def check_dt2_to_dt3(self, dt2, dt3):
    #     self.pp("check dt2 to dt3")
    #     bol = False
    #     for r in tqdm(dt2.itertuples(), disable=self.hideProgressBar):
    #         name_ind = r.name_ind
    #         name = r.name
    #         date_number = r.date_number
    #         var_vals = r.vals
    #         dclass_ind = r.dclass_ind
    #         var_vals_dt3 = dt3.query("name_ind == @name_ind & date_number == @date_number & dclass_ind == @dclass_ind")["vals"].sum()
    #
    #         if var_vals != var_vals_dt3:
    #             bol = True
    #             self.pp("dt2 not eq dt3")
    #             self.pp(var_vals)
    #             self.pp(var_vals_dt3)
    #             status = "dt2 not eq dt3: " + name + ", " + str(date_number)+", dt2:"+str(var_vals)+", dt3:"+str(var_vals_dt3)
    #             tmp_se = pd.Series([self.log_class[3], status, "dt2_not_eq_dt3"],index=self.log_cols)
    #             self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #             self.pp(status)
    #     if bol:
    #         self.pp("dt2 not equal to dt3.")

    # def check_contWork(self, dt1):
    #     self.pp("checking continous working...")
    #     bol = False
    #     for m in tqdm(self.x.df_MemberCapacity["name"], disable=self.hideProgressBar):
    #
    #         # self.pp(m)
    #         m_capa = int(self.x.df_MemberCapacity.query("name == @m")["CapableContinuousWorks"])
    #         max_date = int(dt1["date_number"].max())
    #         # # self.pp(m_capa)
    #         dt1m = dt1.query("name == @m")
    #         #     and date_number < @max_date -@m_capa
    #         for d in range(0, max_date - m_capa + 1):
    #
    #             #         self.pp(d)
    #             # var_tuple = dt1m.query("@d+1 <= date_number <= @d +@m_capa")["var_name"].to_list()
    #             var_1_sum = dt1m.query("@d+1 <= date_number <= @d +@m_capa")["vals"].sum()
    #             if var_1_sum > m_capa:
    #                 bol = True
    #                 status = "continous_work_over: "+ m +" " +str(d+1)+" to "+str(d + m_capa)
    #                 tmp_se = pd.Series([self.log_class[0], status, "over_work"],index=self.log_cols)
    #                 self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #                 self.pp(status)
    #     if bol:
    #         self.pp("over contineous work")

    def check_contNight(self, dt3):
        self.pp("checking continenous nights works...")
        bol = False
        max_date = self.x.max_date
        for m in tqdm(self.x.df_MemberCapacity["name"], disable=self.hideProgressBar):
            # self.pp(m)
            dt3m = dt3.query("name == @m and isNightWork ==1 and vals == 1")
            m_capa = int(self.x.df_MemberCapacity.query(
                "name == @m")["CapableContinuousNightWorks"])

            for d in range(0, max_date - m_capa + 1):
                #         self.pp(d)
                vals_df = dt3m.query(
                    "@d +1 <= date_number <= @d +@m_capa +1").drop_duplicates(subset="date_number")

                if len(vals_df) != 0:
                    vals = vals_df["vals"].sum()
                    if vals == m_capa + 1:
                        bol = True

                        status = "continous_night_work_over: " + m + ", start:" + \
                            str(d + 1) + ", end:" + str(d + m_capa + 1)

                        clps_dict = self.clps_dict_org.copy()
                        clps_dict["members"] = str(m)
                        clps_dict["date_number"] = d + 1
                        clps_dict["class"] = "cnight"
                        clps_dict["info"] = status

                        self.pp(clps_dict)
                        self.clps_list.append(clps_dict)

                        # tmp_se = pd.Series([self.log_class[1], status, "over_night_cont_work"], index=self.log_cols)
                        # self.log_df = self.log_df.append(tmp_se, ignore_index=True)
                        # self.pp(status)
        if bol:
            self.pp("over contineous night work...")

    # def check_contNight(self, dt3, max_count=3):
    #     self.pp("checking continenous nights works...")
    #     bol = False
    #     max_night_cont = max_count
    #     max_date = self.x.max_date
    #     for m in tqdm(self.x.df_MemberCapacity["name"], disable=self.hideProgressBar):
    #         # self.pp(m)
    #         dt3m = dt3.query("name == @m and isNightWork ==1")
    #
    #         for d in range(0, max_date - max_night_cont + 1):
    #             #         self.pp(d)
    #             vals_df = dt2m.query("@d +1 <= date_number <= @d +@max_night_cont")
    #
    #             if len(vals_df) != 0:
    #                 vals = vals_df["vals"].sum()
    #                 if vals == max_count:
    #                     bol = True
    #                     status = "continous_night_work_over: " + m + " " + str(d + 1) + " to " + str(d + max_date +1)
    #                     tmp_se = pd.Series([self.log_class[1], status, "over_night_cont_work"], index=self.log_cols)
    #                     self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #                     self.pp(status)
    #     if bol:
    #         self.pp("over contineous night work...")

    # def check_OnePointOne(self, dt3):
    #     self.pp("checking one task have one person..")
    #     bol = False
    #
    #
    #     dt3_uni = dt3.drop_duplicates(subset=["date_number", "task_num", "dclass_ind","count_in_task"]).loc[:,
    #               ["date_number", "task_num", "dclass_ind", "count_in_task"]]
    #     for ind, date_num, task_num, dclass_ind, count_in_task in tqdm(dt3_uni.itertuples(name=None), disable=self.hideProgressBar):
    #         #     self.pp(date_num, task_num, count_in_task)
    #         val = dt3.query(
    #             "date_number == @date_num and task_num == @task_num and dclass_ind==@dclass_ind and count_in_task == @count_in_task")["vals"].sum()
    #
    #         if val != count_in_task:
    #             bol = True
    #             status = "one_point_one_not_fill: " + str(date_num) + " " + str(task_num) + " " + str(count_in_task) + "_not_eq_" +str(val)
    #             tmp_se = pd.Series([self.log_class[5], status, "one_point_one_not_fill"], index=self.log_cols)
    #             self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #             self.pp(status)
    #     if bol:
    #         self.pp("not one point one...")
    #     else:
    #         self.pp("clear!")

    def check_forbidenSeq(self, dt3):
        self.pp("checking is there forbiden seq")
        bol = False

        df = self.x.df_ForbiddenSequence.query("bef_num !=0 and aft_num !=0")
        fbd_set = self.x.set_ForbiddenSequence_ind

        for m in dt3["name"].unique():
            # self.pp(m)
            dt3m = dt3.query("name ==@m and vals ==1")

            dt3m_ab = dt3m.loc[:, ["date_number", "week_num",
                                   "DutyClass", "dclass_ind", "var_name"]]
            dt3m_ab["after"] = dt3m_ab["date_number"] + 1
            # dt2m_ab
            dt3m_lim = dt3m.loc[:, ["date_number",
                                    "DutyClass", "dclass_ind", "var_name"]]
            dt3m_lim.columns = ["date_number2",
                                "DutyClass2", "dclass_ind2", "var_name2"]
            dt3m_ab2 = pd.merge(dt3m_ab, dt3m_lim, left_on="after",
                                right_on="date_number2", how="inner")

            if len(dt3m_ab2) != 0:
                for r in dt3m_ab2.itertuples():
                    if (r.dclass_ind, r.dclass_ind2) in fbd_set:
                        bol = True
                        status = "forbiden seq was found: " + m + ", s1:" + \
                            str(r.date_number) + str(r.DutyClass) + \
                            ", s2: " + str(r.date_number2) + str(r.DutyClass2)
                        clps_dict = self.clps_dict_org.copy()
                        clps_dict["members"] = str(m)
                        clps_dict["date_number"] = r.date_number
                        clps_dict["class"] = "fseq"
                        clps_dict["info"] = status

                        self.pp(clps_dict)
                        self.clps_list.append(clps_dict)

                        # tmp_se = pd.Series([self.log_class[5], status, "forbiden seq was found"],
                        #                    index=self.log_cols)
                        # self.log_df = self.log_df.append(tmp_se, ignore_index=True)
        if bol:
            self.pp("forbiden seq found!")
        else:
            self.pp("clear!")

    def check_NeedSeq(self, dt3):
        self.pp("checking needed pair is pair...")
        bol = False

        df = self.x.df_NeededSequence.query("bef_num !=0 and aft_num !=0")
        needs_set = self.x.set_NeedSeq_ind
        needs_set_each = set()
        es_1 = set()
        es_2 = set()
        for e in needs_set:
            needs_set_each.add(e[0])
            needs_set_each.add(e[1])
            es_1.add(e[0])
            es_2.add(e[1])

        # その勤務dt3
        dt3m = dt3.query("vals == 1")
        # そのヒトの勤務dt3の日付とか
        dt3m_ab = dt3m.loc[:, ["name", "date_number",
                               "week_num", "DutyClass", "dclass_ind", "var_name"]]
        # 翌日番号
        dt3m_ab["after"] = dt3m_ab["date_number"] + 1

        # 翌日DF
        dt3m_lim = dt3m.loc[:, ["name", "date_number",
                                "DutyClass", "dclass_ind", "var_name"]].copy()
        dt3m_lim.columns = ["name", "date_number2",
                            "DutyClass2", "dclass_ind2", "var_name2"]

        # 翌日番号でマージ
        dt3m_ab2 = pd.merge(dt3m_ab, dt3m_lim, left_on=[
                            "name", "after"], right_on=["name", "date_number2"])
        # 必須DFとマージ
        dt3m_abm = pd.merge(dt3m_ab2, df, left_on=["DutyClass", "DutyClass2"], right_on=["Before", "After"], how="outer") \
            .query("dclass_ind in @es_1 or dclass_ind2 in @es_2")

        dt3m_abm["is_in_needs"] = [1 if (bef, aft) in needs_set else 0 for bef, aft in zip(
            dt3m_abm["dclass_ind"], dt3m_abm["dclass_ind2"])]

        # 逸脱条件
        dt3m_abm = dt3m_abm.query("is_in_needs == 0")

        for r in dt3m_abm.itertuples():
            bol = True

            status = "needed seq collapsed: " + r.name + \
                     ", d1:" + str(r.date_number) + "_" + r.DutyClass +\
                     ", d2:" + str(r.date_number2) + "_" + r.DutyClass2
            clps_dict = self.clps_dict_org.copy()
            clps_dict["members"] = str(r.name)
            clps_dict["date_number"] = r.date_number
            clps_dict["class"] = "nseq"
            clps_dict["info"] = status

            self.pp(clps_dict)
            self.clps_list.append(clps_dict)

        if bol:
            self.pp("needed seq collapsed!")
        else:
            self.pp("clear")
        return dt3m, dt3m_abm, needs_set_each

    def check_ReqNumber(self, dt3):
        bol = False
        self.pp("number is correct?")
        df_req = self.x.df_RequireTask.query("count != 0")
        dt3 = dt3.query("vals==1")
        for r in df_req.itertuples():
            #     self.pp(r)
            r_count = r.count
            r_date = r.date_number
            r_task = r.task_ind

            dt3_r = dt3.query("date_number == @r_date and task_ind == @r_task")

            if len(dt3_r) != r_count:
                bol = True
                status = "collapse req num was found: " + str(r.date_number) + "_" + str(
                    r.task) + " req:" + str(r_count) + ", commit:" + str(len(dt3_r))
                self.pp(status)

                bol = True

                clps_dict = self.clps_dict_org.copy()
                clps_dict["dutys"] = str(r.task)
                clps_dict["date_number"] = r.date_number
                clps_dict["class"] = "rqn"
                clps_dict["info"] = status
                clps_dict["comments"] = str(r.date_number) + "_" + str(
                    r.task) + " req:" + str(r_count) + ", commit:" + str(len(dt3_r))
                self.pp(clps_dict)
                self.clps_list.append(clps_dict)
                # tmp_se = pd.Series([self.log_class[5], status, "not req num was found"],
                #                    index=self.log_cols)
                # self.log_df = self.log_df.append(tmp_se, ignore_index=True)
        if bol:
            self.pp("req_number collapsed!")
        else:
            self.pp("clear")

            # for m in dt2["name"].unique():
            #     # self.pp(m)
            #     df_i = dt2.query("name ==@m and vals ==1")
            #     for date_num in range(1, df_i["date_number"].max()):
            #         # self.pp(date_num)
            #         var_pre = df_i.query("date_number == @date_num")["DutyClass"].to_list()
            #         var_pos = df_i.query("date_number == @date_num +1")["DutyClass"].to_list()
            #         if (len(var_pre) != 0):
            #             if len(var_pos) == 0:
            #                 for t in var_pos:
            #                     if t in self.x.dict_NeedSeq.keys():
            #                         status = "needed seq was not pair: " + m + " " + str(date_num) + " " + str(c)
            #                         self.pp(status)
            #                         tmp_se = pd.Series([self.log_class[5], status, "needed seq was not pair"],
            #                                            index=self.log_cols)
            #                         self.log_df = self.log_df.append(tmp_se, ignore_index=True)
            #             else:
            #                 for c in it.product(var_pre, var_pos):
            #                     # self.pp(c)
            #                     if c[0] in self.x.dict_NeedSeq.keys():
            #                         if c[1] not in self.x.dict_NeedSeq[c[0]]:
            #                             bol = True
            #                             status = "needed seq was not pair: " + m + " " + str(date_num) + " " + str(c)
            #                             self.pp(status)
            #                             tmp_se = pd.Series([self.log_class[5], status, "needed seq was not pair"],
            #                                                index=self.log_cols)
            #                             self.log_df = self.log_df.append(tmp_se, ignore_index=True)

    def check_ReqWorkHour(self, dt3):
        bol = False
        self.pp("is total hour correct?")
        # num_wk_min_proc = self.x.num_wk_min_proc
        # num_wk_max_proc = self.x.num_wk_max_proc
        # max_work_hour = num_wk_max_proc * self.x.total_weekdays
        # min_work_hour = num_wk_min_proc * self.x.total_weekdays

        max_work_hour = self.x.num_wk_max_proc
        min_work_hour = self.x.num_wk_min_proc

        dt3_sum = dt3.query("vals == 1")\
            .groupby("name")["WORK_TIME"]\
            .sum()\
            .reset_index()\
            .query("WORK_TIME < @min_work_hour or @max_work_hour < WORK_TIME")
        if len(dt3_sum) > 0:
            bol = True
            self.pp(dt3_sum)

            for r in dt3_sum.itertuples():
                status = "collapse req hours found:" + r.name +\
                    " min-max:" + str(min_work_hour) + "to"+str(max_work_hour) +\
                    " worked:" + str(r.WORK_TIME)

                clps_dict = self.clps_dict_org.copy()
                clps_dict["name"] = str(r.name)

                clps_dict["class"] = "rqh"
                clps_dict["info"] = status
                clps_dict["comments"] = "min-max:" + str(min_work_hour) + "_to_"+str(max_work_hour) +\
                    " worked:" + str(r.WORK_TIME)

                self.pp(clps_dict)
                self.clps_list.append(clps_dict)

            # tmp_se = pd.Series([self.log_class[5], status, "not req hours was found"],
            #                    index=self.log_cols)
            # self.log_df = self.log_df.append(tmp_se, ignore_index=True)

        if bol:
            self.pp("req_work collapsed!")
        else:
            self.pp("clear")

    def check_TotalWork(self, dt3):
        self.pp("is total work times is correct?")
        bol = False

        ss_days = ["Sunday", "Saturday"]
        len_ss = self.x.df_weeks.query("days in @ss_days")["days"].count()

        len_month = self.x.df_weeks["days"].count()

        max_work_shifts = len_month - len_ss

        dt3 = dt3.copy().query("vals ==1") \
            .groupby("name")["vals"] \
            .sum()
        dt3_check = pd.merge(dt3, self.x.df_MemberCapacity,
                             on="name") \
            .assign(max_work_shifts=max_work_shifts)\
            .assign(CapableTotalWorks=lambda df: [min(i, j) for i, j in zip(df.CapableTotalWorks, df.max_work_shifts)])\
            .query("CapableTotalWorks < vals")

        if len(dt3_check) > 0:
            bol = True
            for r in dt3_check.itertuples():
                status = "collapse max total work was found:" + r.name +\
                    " max:" + str(r.CapableTotalWorks) + ", count:"+str(r.vals)

                clps_dict = self.clps_dict_org.copy()
                clps_dict["name"] = str(r.name)

                clps_dict["class"] = "tt"
                clps_dict["info"] = status

                self.pp(clps_dict)
                self.clps_list.append(clps_dict)
            self.pp(dt3_check)

            # status = "over workers were found"
            # tmp_se = pd.Series([self.log_class[5], status, "overworker was found"],
            #                    index=self.log_cols)
            # self.log_df = self.log_df.append(tmp_se, ignore_index=True)

        if bol:
            self.pp("req_total_work collapsed!")
        else:
            self.pp("clear")

    def check_TotalNights(self, dt3):
        self.pp("is total night is correct?")
        bol = False
        dt3 = dt3.copy().query("vals ==1 and isNightWork ==1") \
            .groupby("name")["WorkHours"] \
            .sum() \
            .reset_index()
        dt3_check = pd.merge(dt3, self.x.df_MemberCapacity, on="name") \
            .query("CapableTotalWorkNightHours < WorkHours")

        if len(dt3_check) > 0:
            bol = True
            # self.pp(dt3_check)

            for r in dt3_check.itertuples():
                status = "collapse max total night work was found:" + r.name +\
                    " max:" + str(r.CapableTotalWorkNightHours) + \
                    ", count:"+str(r.WorkHours)

                clps_dict = self.clps_dict_org.copy()
                clps_dict["name"] = str(r.name)

                clps_dict["class"] = "tn"
                clps_dict["info"] = status

                self.pp(clps_dict)
                self.clps_list.append(clps_dict)
            # self.pp(dt3_check)

        if bol:
            self.pp("req_total_work collapsed!")
        else:
            self.pp("clear")

    def check_NeedSeqFromLast(self, dt3):
        self.pp("is needed seq is preserved?")
        bol = False
        dt3 = dt3.query("date_number == 1")
        dt3_mg = pd.merge(dt3, self.x.df_FinalStatusAtLastMonth, on="name")
        dt3_mg_lim = dt3_mg.loc[:, [
            "name", "LastDay_DutyClass", "DutyClass", "vals"]]
        # dt3_mg_lim
        df_need = self.x.df_NeededSequence.query(
            "bef_num != 0 and aft_num != 0")

        # これも絶対早くできるが。。。
        for r in df_need.itertuples():
            need_bef = r.Before
            need_aft = r.After
            self.pp(need_bef, need_aft)
            dt3_need = dt3_mg_lim.query(
                "LastDay_DutyClass in @need_bef and DutyClass != @need_aft and vals ==1")
            if len(dt3_need) != 0:
                bol = True
                for r in dt3_need.itertuples():

                    status = "collapse in need seq from last detected:" + r.name +\
                        ", last:" + r.LastDay_DutyClass + ", pointed:" + r.DutyClass

                    self.pp(status)
                    self.pp(dt3_need)
                    clps_dict = self.clps_dict_org.copy()
                    clps_dict["name"] = str(r.name)

                    clps_dict["class"] = "lnd"
                    clps_dict["info"] = status

                    self.pp(clps_dict)
                    self.clps_list.append(clps_dict)
                # tmp_se = pd.Series([self.log_class[5], status, "co in ned seq frm lst was found"],
                #                   index=self.log_cols)
                # self.log_df = self.log_df.append(tmp_se, ignore_index=True)
        if bol:
            self.pp("collapse in need seq!")
        else:
            self.pp("clear")

    def check_ForbidSeqFromLast(self, dt3):
        self.pp("is forbid seq is avoided?")
        bol = False
        dt3 = dt3.query("date_number == 1")
        dt3_mg = pd.merge(dt3, self.x.df_FinalStatusAtLastMonth, on="name")
        dt3_mg_lim = dt3_mg.loc[:, [
            "name", "LastDay_DutyClass", "DutyClass", "vals"]]
        df_fbd = self.x.df_ForbiddenSequence.query(
            "bef_num != 0 and aft_num != 0")
        # df_fbd
        for r in df_fbd.itertuples():
            need_bef = r.Before
            need_aft = r.After
            self.pp(need_bef, need_aft)
            dt3_fbd = dt3_mg_lim.query(
                "LastDay_DutyClass in @need_bef and DutyClass == @need_aft and vals ==1")
            if len(dt3_fbd) != 0:
                bol = True

                self.pp(dt3_fbd)
                for r in dt3_fbd.itertuples():
                    status = "collapse in fbd seq from last detected:" + r.name +\
                        ", last:" + r.LastDay_DutyClass + ", pointed:" + r.DutyClass

                    # self.pp(status)

                    clps_dict = self.clps_dict_org.copy()
                    clps_dict["name"] = str(r.name)

                    clps_dict["class"] = "lfbd"
                    clps_dict["info"] = status

                    self.pp(clps_dict)
                    self.clps_list.append(clps_dict)
                # tmp_se = pd.Series([self.log_class[5], status, "colps in fbd frm lst was found"],
                #                    index=self.log_cols)
                # self.log_df = self.log_df.append(tmp_se, ignore_index=True)
        if bol:
            self.pp("collapse in fbd seq!")
        else:
            self.pp("clear")

    def check_ContWork_FromLast(self, dt1):
        bol = False
        self.pp("checking work count from last month...")
        dt1 = dt1.copy()
        for m in dt1["name"].unique():
            dt1_s = dt1.query("name == @m")

            # min zeroは要するに最初のoff日
            min_zero = dt1_s.query("vals == 0")["date_number"].min()

            # max 1stは要するに最大連勤-最終先月連勤
            max_1st = self.x.df_CapaAndLast.query(
                "name == @m")["max_1st_cont"].tolist()[0]
            if min_zero > max_1st+1 and min_zero > 1:
                # 最初のオフ日が、最大継続連勤+1より大きいとだめ
                # 初日がオフならそもそも問題ない
                bol = True
                status = "over cont work from last month:" + m + ", 1st cont work:" + \
                    str(min_zero) + ", last cont work:" + str(max_1st)
                self.pp(status)
                clps_dict = self.clps_dict_org.copy()
                clps_dict["name"] = str(m)

                clps_dict["class"] = "lcnt"
                clps_dict["info"] = status

                self.pp(clps_dict)
                self.clps_list.append(clps_dict)

                # tmp_se = pd.Series([self.log_class[15], status, "over 1st cont work"],
                #                    index=self.log_cols)
                # self.log_df = self.log_df.append(tmp_se, ignore_index=True)
        if bol:
            self.pp("over work from last!")
        else:
            self.pp("clear")

    def check_ContNight_FromLast(self, dt2):
        bol = False
        self.pp("checking night work count from last month")
        dt2 = dt2.copy()
        for m in dt2["name"].unique():
            dt2_s = dt2.query("name == @m and isNightWork == True")
            min_zero = dt2_s.query("vals == 0")["date_number"].min()
            max_1st_night = self.x.df_CapaAndLast.query(
                "name == @m")["max_1st_cont_night"].tolist()[0]
            if min_zero > max_1st_night and min_zero > 1:
                bol = True
                status = "over cont work night from last month:" + m + " 1st cont night work:" + str(
                    min_zero) + ", last cont night work:" + str(max_1st_night)
                self.pp(status)
                clps_dict = self.clps_dict_org.copy()
                clps_dict["name"] = str(m)

                clps_dict["class"] = "lcnn"
                clps_dict["info"] = status

                self.pp(clps_dict)
                self.clps_list.append(clps_dict)
                # tmp_se = pd.Series([self.log_class[16], status, "over 1st cont night work"],
                #                    index=self.log_cols)
                # self.log_df = self.log_df.append(tmp_se, ignore_index=True)
        if bol:
            self.pp("over night work from last!")
        else:
            self.pp("clear")

    def check_ShiftCount(self, dt2):
        bol = False
        self.pp("checking shift count satisfication")
        dt2 = dt2.copy()\
            .query("vals==1")\
            .groupby(["name", "DutyClass"])["DutyClass"] \
            .count() \
            .reset_index(drop=False, name="take_count")

        memcap = self.x.df_MemberCapacity.copy()

        cols = memcap.columns.to_list()
        shifts = self.x.df_DutyClass["DutyClass"].unique().tolist()
        cols = [c for c in cols if c.replace("Capable", "") in shifts]

        memcap = memcap \
            .set_index("name")\
            .loc[:, cols].copy()\
            .unstack()\
            .reset_index()\
            .rename(columns={"level_0": "DutyClass",
                             0: "capable_count"})
        memcap["DutyClass"] = memcap["DutyClass"].str.replace("Capable", "")

        sc_mem_cap = pd.merge(memcap, dt2, on=["name", "DutyClass"], how="left")\
            .fillna(0)\
            .query("capable_count < take_count")

        if len(sc_mem_cap) > 0:
            bol = True
            sc_mem_cap["outs"] = sc_mem_cap["name"] \
                + "_"\
                + sc_mem_cap["DutyClass"] \
                + "_"\
                + sc_mem_cap["capable_count"].astype(str) \
                + "_"\
                + sc_mem_cap["take_count"].astype(str)

            for r in sc_mem_cap.itertuples():
                status = "collapse in shift capa:" + r.name +\
                    ", DutyClass:" + r.DutyClass + \
                    ", capa:" + str(r.capable_count) + \
                    ", taken:" + str(r.take_count)

                # self.pp(status)

                clps_dict = self.clps_dict_org.copy()
                clps_dict["name"] = str(r.name)

                clps_dict["class"] = "sc"
                clps_dict["info"] = status
                clps_dict["comments"] = "DC:" + r.DutyClass + \
                    ", capa:" + str(r.capable_count) + \
                    ", taken:" + str(r.take_count)

                self.pp(clps_dict)
                self.clps_list.append(clps_dict)

            # status = "shift capa over" + ";".join(sc_mem_cap["outs"].to_list())
            # self.pp(status)
            # tmp_se = pd.Series([self.log_class[17], status, "shift cap over"],
            #                     index=self.log_cols)
            # self.log_df = self.log_df.append(tmp_se, ignore_index=True)
        if bol:
            self.pp("capa over!")
        else:
            self.pp("clear")

    def check_MemberPairShift(self, dt3):
        self.pp("checking forbidden pair in shifts")
        bol = False
        dt3m = dt3.query("vals == 1")
        df_fbd = self.x.df_ForbiddenPairShift.copy()
        for r in df_fbd.itertuples():
            n1 = r.name1
            n2 = r.name2
            s = r.shift
            # self.pp(n1, n2, s)
            for dn in dt3m["date_number"].unique():
                # self.pp(dn)
                dt3_dn = dt3m.query("date_number == @dn and DutyClass == @s")\
                    .query("name == @n1 or name == @n2")
                # ここが2以上というのは、要するに禁断の二人が勤務しているということ。1であれば問題ない。
                if len(dt3_dn) > 1:
                    for r in dt3_dn.itertuples():
                        bol = True
                        status = "forbidden member pair shift detected: " + n1 + \
                            "_" + n2+" on shift:"+s+" in data_number:"+str(dn)

                        clps_dict = self.clps_dict_org.copy()
                        clps_dict["name"] = str(r.name)

                        clps_dict["class"] = "mps"
                        clps_dict["info"] = status

                        self.pp(clps_dict)
                        self.clps_list.append(clps_dict)
                        # self.pp(status)
                        # tmp_se = pd.Series([self.log_class[18],status, "forbidden member pair"],
                        #                     index=self.log_cols)
                        # self.log_df = self.log_df.append(tmp_se, ignore_index = True)

        if bol:
            self.pp("fobd member pair shift")
        else:
            self.pp("clear")

    def check_DayABVetran(self, dt3):
        self.pp("checking daytime veteran groups")
        bol = False
        dt3 = dt3.query("vals == 1")
        As = ["A", "A/B"]
        Bs = ["B", "A/B"]
        dfw = self.x.df_week_task.query("tasks == 'ﾆB_S'").copy()
        for d in dt3["date_number"].unique():
            # print(d)
            n_vet_total = dfw\
                .query("date_number == @d")["count_in_task"].to_list()[0]
            n_min_vet_group = int(n_vet_total/2)
            if n_min_vet_group > 1:
                df_a = dt3\
                    .query("date_number == @d")\
                    .query("DutyLabel == 'ﾆB_S'")\
                    .query("group in @As")
                df_b = dt3\
                    .query("date_number == @d")\
                    .query("DutyLabel == 'ﾆB_S'")\
                    .query("group in @Bs")

                #print(len(df_a), len(df_b))
                if not (n_min_vet_group <= len(df_a) <= n_vet_total and n_min_vet_group <= len(df_b) <= n_vet_total):
                    bol = True
                    status = "daytime veteran colapsion: date_" + \
                        str(d) + ", A:" + str(len(df_a)) + \
                        ", B:" + str(len(df_b))
                    clps_dict = self.clps_dict_org.copy()

                    clps_dict["class"] = "dab"
                    clps_dict["info"] = status

                    self.pp(clps_dict)
                    self.clps_list.append(clps_dict)
                    print(status)

        if bol:
            self.pp("daytime veteran groups colapsion")
        else:
            self.pp("clear")

    def check_NightTimeNovice(self, dt3):
        self.pp("checking night time novice groups")
        bol = False
        As = ["A", "A/B"]
        Bs = ["B", "A/B"]
        df3 = dt3.query("vals ==1")
        tag_nights = ["MY_I", "YY_I"]
        for t in tag_nights:
            # print(t)
            for d in df3["date_number"].unique():
                # print(d)
                df_a = df3\
                    .query("date_number == @d")\
                    .query("DutyLabel == @t")\
                    .query("group in @As")
                df_b = df3\
                    .query("date_number == @d")\
                    .query("DutyLabel == @t")\
                    .query("group in @Bs")

                #print(len(df_a), len(df_b))
                if not (len(df_a) <= 1 and len(df_b) <= 1):
                    bol = True
                    status = "night time novice colapsion: date_" + \
                        str(d) + ", A:" + str(len(df_a)) + \
                        ", B:" + str(len(df_b))
                    clps_dict = self.clps_dict_org.copy()

                    clps_dict["class"] = "nnv"
                    clps_dict["info"] = status

                    self.pp(clps_dict)
                    self.clps_list.append(clps_dict)
                    # print(status)
        if bol:
            self.pp("night time novice colapsion")
        else:
            self.pp("clear")

    def tags_in(self, tag):
        return (len(self.exclude_list) == 0 and tag in self.include_list) or (len(self.include_list) == 0 and tag not in self.exclude_list)

    def checkAllLimitation(self, include=[], exclude=[], res_parser=None, hideProgressBar=False):
        self.pp("checking!")

        rp = res_parser
        self.hideProgressBar = hideProgressBar

        self.hideProgressBar = hideProgressBar
        self.include_list = include
        self.exclude_list = exclude

        tag = "1by1"
        if self.tags_in(tag):
            self.check_one_by_one_task(dt3=rp.dt3_mg)

        tag = "cwork"
        if self.tags_in(tag):
            self.check_contWork(dt3=rp.dt3_mg)

        tag = "cnight"
        if self.tags_in(tag):
            self.check_contNight(dt3=rp.dt3_mg)

        tag = "fseq"
        if self.tags_in(tag):
            self.check_forbidenSeq(dt3=rp.dt3_mg)

        tag = "nseq"
        if self.tags_in(tag):
            self.check_NeedSeq(dt3=rp.dt3_mg)

        # equi groupの評価関数はどのような定量評価が良いか思案中なので無い

        tag = "rqn"
        if self.tags_in(tag):
            self.check_ReqNumber(dt3=rp.dt3_mg)

        tag = "rqh"
        if self.tags_in(tag):
            self.check_ReqWorkHour(dt3=rp.dt3_mg)

        tag = "tt"
        if self.tags_in(tag):
            self.check_TotalWork(dt3=rp.dt3_mg)

        tag = "tn"
        if self.tags_in(tag):
            self.check_TotalNights(dt3=rp.dt3_mg)

        tag = "lfbd"
        if self.tags_in(tag):
            self.check_ForbidSeqFromLast(dt3=rp.dt3_mg)

        tag = "lnd"
        if self.tags_in(tag):
            self.check_NeedSeqFromLast(dt3=rp.dt3_mg)

        tag = "lcnt"
        if self.tags_in(tag):
            self.check_ContWork_FromLast(dt1=rp.dt1_mg)

        tag = "lcnn"
        if self.tags_in(tag):
            self.check_ContNight_FromLast(dt2=rp.dt2_mg)

        tag = "sc"
        if self.tags_in(tag):
            self.check_ShiftCount(dt2=rp.dt2_mg)

        tag = "mps"
        if self.tags_in(tag):
            self.check_MemberPairShift(dt3=rp.dt3_mg)

        tag = "dab"
        if self.tags_in(tag):
            self.check_DayABVetran(dt3=rp.dt3_mg)

        tag = "nnv"
        if self.tags_in(tag):
            self.check_NightTimeNovice(dt3=rp.dt3_mg)

        # all collapse sum
        self.log_df = pd.DataFrame(self.clps_list)

    #     したの古いコメントアウト関数は、実現値引くキャパからの差分を等しくなれぞかし、というチェックをしていた。
    #     インプットはdt1であった。
    #     print("checking date equality...")
    #     dt = dt1.copy().query("vals ==1").loc[:, ["name", "vals"]].groupby("name").sum()
    #
    #     df2 = pd.merge(dt, self.x.df_MemberCapacity, on="name")
    #     df2["total_work_sub"] = df2["CapableTotalWorks"] - df2["vals"]
    #     v=df2["total_work_sub"].var()
    #     if v >= var_cutoff:
    #
    #
    #         bol = True
    #         status = "works dist over var: " + str(v)
    #         print(status)
    #
    #         tmp_se = pd.Series([self.log_class[8], status, "uneq works"],
    #                            index=self.log_cols)
    #         self.log_df = self.log_df.append(tmp_se, ignore_index=True)

    # def check_NeedSeq(self, dt2):
    #     print("checking needed pair is pair...")
    #     bol = False
    #     for m in dt2["name"].unique():
    #         # print(m)
    #         df_i = dt2.query("name ==@m and vals ==1")
    #         for date_num in range(1, df_i["date_number"].max()):
    #             # print(date_num)
    #             var_pre = df_i.query("date_number == @date_num")["DutyClass"].to_list()
    #             var_pos = df_i.query("date_number == @date_num +1")["DutyClass"].to_list()
    #             if (len(var_pre) != 0) :
    #                 if len(var_pos) == 0:
    #                     for t in var_pos:
    #                         if t in self.x.dict_NeedSeq.keys():
    #                             status = "needed seq was not pair: " + m + " " + str(date_num) + " " + str(c)
    #                             print(status)
    #                             tmp_se = pd.Series([self.log_class[5], status, "needed seq was not pair"],
    #                                                index=self.log_cols)
    #                             self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #                 else:
    #                     for c in it.product(var_pre, var_pos):
    #                         # print(c)
    #                         if c[0] in self.x.dict_NeedSeq.keys():
    #                             if c[1] not in self.x.dict_NeedSeq[c[0]]:
    #                                 bol = True
    #                                 status = "needed seq was not pair: " +m +" "+str(date_num) + " " + str(c)
    #                                 print(status)
    #                                 tmp_se = pd.Series([self.log_class[5], status, "needed seq was not pair"],
    #                                                    index=self.log_cols)
    #                                 self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #     if bol:
    #         print("needed seq...")
    #
    # def check_MemberAbility(self, dt3):
    #     print("checking over ability..")
    #     dt = dt3.copy()
    #     dt = dt.query("ability ==0 and vals ==1")
    #
    #     dt.loc[:,"status"] = "abily_over: " + dt["name"] + "_" + dt["date_number"].astype(str) + "_" + dt["DutyClass"] + "_" + \
    #                    dt["DutyLabel"]
    #     dt.loc[:,"comment"] = "ability over"
    #     dt.loc[:,"log_class"] = self.log_class[6]
    #     dt = dt.loc[:,["log_class", "status", "comment"]].copy()
    #     # print(dt)
    #     dt.columns = self.log_cols
    #     self.log_df.append(dt)
    #
    # def check_WantedRest(self, dt3):
    #     print("checking wanted rest ability..")
    #     dt = dt3.copy()
    #     dt = dt.query("takability ==0 and vals ==1")
    #
    #     dt.loc[:,"status"] = "takability_over: " + dt["name"] + "_" + dt["date_number"].astype(str) + "_" + dt["DutyClass"] + "_" + \
    #                    dt["DutyLabel"]
    #     dt.loc[:,"comment"] = "takability over"
    #     dt.loc[:,"log_class"] = self.log_class[7]
    #     dt = dt.loc[:, ["log_class", "status", "comment"]].copy()
    #     # print(dt)
    #     dt.columns = self.log_cols
    #     self.log_df.append(dt)
    #
    # def check_DateEq(self, dt1, var_cutoff = 3):
    #     bol = False
    #     print("checking date equality...")
    #     dt = dt1.copy().query("vals ==1").loc[:, ["name", "vals"]].groupby("name").sum()
    #
    #     df2 = pd.merge(dt, self.x.df_MemberCapacity, on="name")
    #     df2["total_work_sub"] = df2["CapableTotalWorks"] - df2["vals"]
    #     v=df2["total_work_sub"].var()
    #     if v >= var_cutoff:
    #
    #
    #         bol = True
    #         status = "works dist over var: " + str(v)
    #         print(status)
    #
    #         tmp_se = pd.Series([self.log_class[8], status, "uneq works"],
    #                            index=self.log_cols)
    #         self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #
    # def check_NightEq(self, dt2, var_cutoff=3):
    #     bol = False
    #     print("checking night equality...")
    #     dt = dt2.copy().query("vals ==1 and isNightWork ==1").loc[:, ["name", "vals"]].groupby("name").sum()
    #
    #     df2 = pd.merge(dt, self.x.df_MemberCapacity, on="name")
    #     df2["night_work_sub"] = df2["CapableNights"] - df2["vals"]
    #     v = df2["night_work_sub"].var()
    #     if v >= var_cutoff:
    #         bol = True
    #         status = "night works dist over var: " + str(v)
    #         print(status)
    #
    #         tmp_se = pd.Series([self.log_class[11], status, "uneq night works"],
    #                            index=self.log_cols)
    #         self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #
    # def check_OverCapTotalWorks(self,dt1):
    #     bol = False
    #     print("checking date over...")
    #     dt = dt1.copy().query("vals ==1").loc[:, ["name", "vals"]].groupby("name").sum()
    #
    #     df2 = pd.merge(dt, self.x.df_MemberCapacity, on="name")
    #     df2["total_work_sub"] = df2["CapableTotalWorks"] - df2["vals"]
    #     df3 = df2.query("total_work_sub < 0")
    #     # print(df3)
    #     if len(df3) != 0:
    #         bol = True
    #         print("total works over!")
    #         df3["status"] = df3["name"] + "_" + df3["total_work_sub"].astype(str) +"_"+df3["CapableTotalWorks"].astype(str)
    #         df3["comment"] = "total_works_over"
    #         df3["log_class"] = self.log_class[9]
    #         df4 = df3.loc[:,["log_class", "status","comment"]]
    #         # print(df4)
    #         self.log_df.append(df4)
    #
    #
    # def check_OverCapNightWorks(self, dt2):
    #     bol = False
    #     print("checking night equality")
    #     df = dt2.copy()
    #     df = df.query("isNightWork ==1 and vals ==1")[["name", "vals"]].copy().groupby("name").sum()
    #     df2 = pd.merge(df, self.x.df_MemberCapacity, on="name")
    #     df2["night_work_sub"] = df2["CapableNights"] - df2["vals"]
    #     df3 = df2.query("night_work_sub < 0")
    #     if len(df3) != 0:
    #         bol = True
    #         print("total night over!")
    #         df3.loc[:,"status"] = df3["name"] + "_" + df3["night_work_sub"].astype(str) + "_" + df3[
    #             "CapableNights"].astype(str)
    #         df3.loc[:,"comment"] = "night_works_over"
    #         df3.loc[:,"log_class"] = self.log_class[10]
    #         df4 = df3.loc[:, ["log_class", "status", "comment"]]
    #         # print(df4)
    #         self.log_df.append(df4)
    #
    # def check_GroupEqDT3(self, dt3, cut_off = 1):
    #     bol=False
    #     print("checking group equality..")
    #     df = dt3.copy()
    #     df = df.query("count_in_task != 1 and vals != 0")
    #     df_lim = df[
    #         ["date_number", "DutyClass", "DutyLabel", "dclass_ind", "task_ind", "count_in_task"]].drop_duplicates()
    #     df_a = df.query("group =='A'")
    #     df_b = df.query("group =='B'")
    #
    #     l_len = []
    #     for r in df_lim.itertuples():
    #         date_number = r.date_number
    #         dclass_ind = r.dclass_ind
    #         task_ind = r.task_ind
    #
    #         a = df_a.query("date_number == @date_number and dclass_ind == @dclass_ind and task_ind == @task_ind")
    #         b = df_b.query("date_number == @date_number and dclass_ind == @dclass_ind and task_ind == @task_ind")
    #         l_len.append(len(a) - len(b))
    #
    #     df_lim["ab_sub"] = l_len
    #     v = df_lim["ab_sub"].var()
    #     if v > cut_off:
    #         print("group dist over in dt3...")
    #         status = "each group dist over var at dt2: " + str(v)
    #         print(status)
    #
    #         tmp_se = pd.Series([self.log_class[12], status, "dt3 over dist"],
    #                            index=self.log_cols)
    #         self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #
    # def check_GroupEqDT2(self, dt2, cut_off=1):
    #     bol=False
    #     print("checking group equality at dt2")
    #     df = dt2.copy()
    #     df = df.query("vals != 0")
    #     df_lim = df[
    #         ["date_number", "DutyClass", "dclass_ind"]].drop_duplicates()
    #     df_a = df.query("group =='A'")
    #     df_b = df.query("group =='B'")
    #
    #     l_len = []
    #     for r in df_lim.itertuples():
    #         date_number = r.date_number
    #         dclass_ind = r.dclass_ind
    #
    #         a = df_a.query("date_number == @date_number and dclass_ind == @dclass_ind")
    #         b = df_b.query("date_number == @date_number and dclass_ind == @dclass_ind")
    #         l_len.append(len(a) - len(b))
    #
    #     df_lim["ab_sub"] = l_len
    #     v = df_lim["ab_sub"].var()
    #     if v > cut_off:
    #         print("group dist over in dt2...")
    #         status = "each group dist over var at dt2: " + str(v)
    #         print(status)
    #
    #         tmp_se = pd.Series([self.log_class[12], status, "dt2 over dist"],
    #                            index=self.log_cols)
    #         self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #
    # def check_ForbidSeq_FromLast(self,dt2):
    #     print("checking forbiden seq from last month...")
    #     dt2 = dt2.query("date_number==1 and vals ==1")
    #     dt2x = pd.merge(self.x.df_FinalStatusAtLastMonth, dt2, on="name")
    #     # dt2x
    #     for r in dt2x.itertuples():
    #         if (r.LastDay_DutyClass, r.DutyClass) in self.x.set_ForbiddenSequence:
    #             status = "forbiden seq in 1st day:" + r.name + " " + r.LastDay_DutyClass + "_" + r.DutyClass
    #             print(status)
    #             tmp_se = pd.Series([self.log_class[13], status, "forbid sed from last"],
    #                                index=self.log_cols)
    #             self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #
    # def check_NeedPair_FromLast(self,dt2):
    #     print("checking needed pair from last month...")
    #     need_before = set()
    #     dict_needs = dict()
    #     for i in self.x.set_NeedSeq:
    #         need_before.add(i[0])
    #         dict_needs[i[0]] = i[1]
    #     # print(need_before)
    #     dt2 = dt2.copy()
    #     dt2x = pd.merge(self.x.df_FinalStatusAtLastMonth, dt2, on="name").query("date_number==1 and LastDay_DutyClass in @need_before")
    #     dt2x_lim = dt2x.drop_duplicates(subset=["name", "LastDay_DutyClass"])
    #
    #     for r in dt2x.drop_duplicates(subset=["name", "LastDay_DutyClass"]).itertuples():
    #         dt2x_sub = dt2x.query("name == @r.name and LastDay_DutyClass == @r.LastDay_DutyClass and vals==1")
    #         if len(dt2x_sub) == 0 or dt2x_sub["DutyClass"].tolist()[0] != dict_needs[r.LastDay_DutyClass]:
    #             status = "needed seq not in 1st day:" + r.name + " " + r.LastDay_DutyClass
    #             print(status)
    #             tmp_se = pd.Series([self.log_class[14], status, "need seq from last"],
    #                                index=self.log_cols)
    #             self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #
    # def check_ContWork_FromLast(self, dt1):
    #     print("checking work count from last month...")
    #     dt1 = dt1.copy()
    #     for m in dt1["name"].unique():
    #         dt1_s = dt1.query("name == @m")
    #         min_zero = dt1_s.query("vals == 0")["date_number"].min()
    #         max_1st = self.x.df_CapaAndLast.query("name == @m")["max_1st_cont"].tolist()[0]
    #         if min_zero > max_1st and min_zero != 0:
    #             status = "over cont work from last month:" + m + " 1st cont work" + str(min_zero) + ", last cont work:" +str(max_1st)
    #             print(status)
    #             tmp_se = pd.Series([self.log_class[15], status, "over 1st cont work"],
    #                                index=self.log_cols)
    #             self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #
    # def check_ContNight_FromLast(self, dt2):
    #     print("checking night work count from last month")
    #     dt2 = dt2.copy()
    #     for m in dt2["name"].unique():
    #         dt2_s = dt2.query("name == @m and isNightWork == True")
    #         min_zero = dt2_s.query("vals == 0")["date_number"].min()
    #         max_1st_night = self.x.df_CapaAndLast.query("name == @m")["max_1st_cont_night"].tolist()[0]
    #         if min_zero > max_1st_night and min_zero != 0:
    #             status = "over cont work night from last month:" + m + " 1st cont night work" + str(
    #                 min_zero) + ", last cont night work:" + str(max_1st_night)
    #             print(status)
    #             tmp_se = pd.Series([self.log_class[16], status, "over 1st cont night work"],
    #                                index=self.log_cols)
    #             self.log_df = self.log_df.append(tmp_se, ignore_index=True)
    #
    # def checkAllLimitation(self, include=[], exclude=[], res_parser=None,hideProgressBar = False):
    #
    #     rp =res_parser
    #     self.hideProgressBar = hideProgressBar

        # if (len(exclude) == 0 and "3dts" in include) or (len(include) == 0 and "3dts" not in exclude):
        #     self.check_dt1_to_dt2(dt1=rp.dt1_mg, dt2=rp.dt2_mg)
        #     self.check_dt2_to_dt3(dt2=rp.dt2_mg, dt3=rp.dt3_mg)
        #
        # if (len(exclude) == 0 and "1p1" in include) or (len(include) == 0 and "1p1" not in exclude):
        #     self.check_OnePointOne(dt3=rp.dt3_mg)
        # # #
        # if (len(exclude) == 0 and "cnight" in include) or (len(include) == 0 and "cnight" not in exclude):
        #     self.check_contNight(dt2=rp.dt2_mg)
        #
        # if (len(exclude) == 0 and "cwork" in include) or (len(include) == 0 and "cwork" not in exclude):
        #     self.check_contWork(dt1=rp.dt1_mg)
        #
        # if (len(exclude) == 0 and "fseq" in include) or (len(include) == 0 and "fseq" not in exclude):
        #     self.check_forbidenSeq(dt2=rp.dt2_mg)
        #
        # if (len(exclude) == 0 and "nseq" in include) or (len(include) == 0 and "nseq" not in exclude):
        #     self.check_NeedSeq(dt2=rp.dt2_mg)
        # #
        # if (len(exclude) == 0 and "oAb" in include) or (len(include) == 0 and "oAb" not in exclude):
        #     self.check_MemberAbility(dt3=rp.dt3_mg)
        #
        # if (len(exclude) == 0 and "wres" in include) or (len(include) == 0 and "wres" not in exclude):
        #     self.check_WantedRest(dt3=rp.dt3_mg)
        #
        # if (len(exclude) == 0 and "eq" in include) or (len(include) == 0 and "eq" not in exclude):
        #     self.check_DateEq(dt1=rp.dt1_mg)
        #     self.check_OverCapTotalWorks(dt1=rp.dt1_mg)
        #
        # if (len(exclude) == 0 and "eqn" in include) or (len(include) == 0 and "eqn" not in exclude):
        #     self.check_NightEq(dt2=rp.dt2_mg)
        #     self.check_OverCapNightWorks(dt2=rp.dt2_mg)
        #     # self.limit_EquiNights()
        #
        # if (len(exclude) == 0 and "eqg3" in include) or (len(include) == 0 and "eqg3" not in exclude):
        #     self.check_GroupEqDT3(rp.dt3_mg)
        #
        # if (len(exclude) == 0 and "eq2" in include) or (len(include) == 0 and "eqg2" not in exclude):
        #     self.check_GroupEqDT2(rp.dt2_mg)
        #
        # if (len(exclude) == 0 and "lfbd" in include) or (len(include) == 0 and "lfbd" not in exclude):
        #     self.check_ForbidSeq_FromLast(dt2=rp.dt2_mg)
        #
        # if (len(exclude) == 0 and "lnd" in include) or (len(include) == 0 and "lnd" not in exclude):
        #     self.check_NeedPair_FromLast(dt2=rp.dt2_mg)
        #
        # if (len(exclude) == 0 and "lcnt" in include) or (len(include) == 0 and "lcnt" not in exclude):
        #     self.check_ContWork_FromLast(dt1=rp.dt1_mg)
        #
        # if (len(exclude) == 0 and "lcnn" in include) or (len(include) == 0 and "lcnn" not in exclude):
        #     self.check_ContNight_FromLast(dt2=rp.dt2_mg)
    # def check_


class table_maker():

    def __init__(self, x, res, hideProgressBar=False):
        print("table_maker activated.")
        self.rp = res_parser(x=x)
        self.rp.prepRes(res=res, hideProgressBar=hideProgressBar)
        self.hideProgressBar = hideProgressBar

        # self.dt1_pos = self.rp.dt1_mg.query("vals == 1")
        self.dt1_pos = self.rp.dt1_mg.copy()
        # self.dt2_pos = self.rp.dt2_mg.query("vals == 1")
        self.dt2_pos = self.rp.dt2_mg.copy()
        # self.dt3_pos = self.rp.dt3_mg.query("vals == 1")
        self.dt3_pos = self.rp.dt3_mg.copy()

        self.datetable = self.getDateTable()
        self.dutytable = self.getDutyTable()
        self.shifttable = self.getShiftTable()
        self.eachtask = self.getEachTaskTable()
        self.member_stat = self.getMemberStats()
        self.teamstats_byshift = self.getTeamStats_byshift()
        self.teamstats_byduty = self.getTeamStats_byduty()

        self.finalstat = self.getFinalState()

    def saveEach(self, prefix="result_table_"):
        self.datetable.to_csv(prefix+"date.csv", sep=",",
                              encoding='utf_8_sig')
        self.dutytable.to_csv(prefix+"duty.csv", sep=",",
                              encoding='utf_8_sig')
        self.shifttable.to_csv(
            prefix+"shift.csv", sep=",", encoding='utf_8_sig')
        self.eachtask.to_csv(prefix+"eachtask.csv",
                             sep=",", encoding='utf_8_sig')
        self.member_stat.to_csv(
            prefix + "member_stats.csv", sep=",", encoding='utf_8_sig')

    def getDateTable(self):
        t1 = self.rp.dt1_mg.copy() \
            .loc[:, ["date_number", "num_days", "days4proc", "name", "vals"]] \
            .drop_duplicates() \
            .set_index(["date_number", "num_days", "days4proc"])
        t1_wide = pd.DataFrame()
        for name in t1["name"].unique():
            # print(name)
            t1_wide[name] = t1.query("name == @name")["vals"]
        return t1_wide.T

    def getShiftTable(self):
        t2 = self.rp.dt2_mg.copy()
        t2 = t2.loc[:, ["name", "date_number", "num_days",
                        "DutyClass", "vals"]].drop_duplicates()
        task_str = [r.DutyClass if r.vals ==
                    1 else "" for r in t2.itertuples()]
        t2["task_str"] = task_str
        t2_lim = t2.loc[:, ["name", "date_number",
                            "num_days"]].drop_duplicates()
        task_strs = []

        for r in t2_lim.itertuples():
            r_name = r.name
            r_date_number = r.date_number
            r_num_days = r.num_days
            l_str = t2.query("name == @r_name and date_number == @r_date_number and num_days == @r_num_days")[
                "task_str"].to_list()
            # print(l_str)
            task_strs.append(self.comb_str(l_str))

        t2_lim["task_str"] = task_strs
        t2_lim = t2_lim.set_index(["date_number", "num_days"])
        t2_wide = pd.DataFrame()
        # t2_wide
        for name in tqdm(t2_lim["name"].unique(), disable=self.hideProgressBar):
            t2_wide[name] = t2_lim.query("name == @name")["task_str"]

        df_Final_sub = self.rp.x.df_FinalStatusAtLastMonth.copy(
        ).loc[:, ["name", "LastDay_DutyClass"]]
        df_Final_sub.index = df_Final_sub["name"]
        df_Final_sub = df_Final_sub.drop("name", axis=1)
        df_Final_sub = df_Final_sub.fillna("")
        idx = pd.MultiIndex.from_arrays(
            [[0], ["last_day"]], names=("date_number", "num_days"))
        df_Final_sub.columns = idx

        dt2x = pd.merge(df_Final_sub, t2_wide.T,
                        left_index=True, right_index=True)
        # dt2x
        return dt2x

    def comb_str(self, l_str):
        l_str_pos = [i for i in l_str if i != ""]
        if len(l_str_pos) == 0:
            return ""
        else:
            return "_".join(l_str_pos)

    def getDutyTable(self):

        t3 = self.rp.dt3_mg.copy()
        # t3
        task_str = [r.DutyLabel if r.vals ==
                    1 else "" for r in t3.itertuples()]
        t3["task_str"] = task_str
        t3_lim = t3.loc[:, ["name", "date_number",
                            "num_days"]].drop_duplicates()
        duty_strs = []

        for r in t3_lim.itertuples():
            r_name = r.name
            r_date_number = r.date_number
            r_num_days = r.num_days
            l_str = t3.query("name == @r_name and date_number == @r_date_number and num_days == @r_num_days")[
                "task_str"].to_list()
            duty_strs.append(self.comb_str(l_str))
        t3_lim["duty_str"] = duty_strs
        t3_wid = t3_lim.pivot_table(values=['duty_str'], index=['date_number', "num_days"], columns=['name'],aggfunc=lambda x: ' '.join(x)).T\
            .reset_index()\
            .drop(columns ="level_0")
            # \
            # .set_index("name")
        # t3_lim = t3_lim.set_index(["date_number", "num_days"])
        # t3_wide = pd.DataFrame()

        # for name in t3_lim["name"].unique():
        #     t3_wide[name] = t3_lim.query("name == @name")["duty_str"]

        df_Final_sub = self.rp.x.df_FinalStatusAtLastMonth.copy(
        ).loc[:, ["name", "LastDay_DutyLabel"]]
        df_Final_sub.index = df_Final_sub["name"]
        df_Final_sub = df_Final_sub.drop("name", axis=1) 
        df_Final_sub = df_Final_sub.fillna("")
        df_Final_sub.columns = [(0, df_Final_sub.columns[0])]
        # df_Final_sub.columns = [("name"), (0, "LastDay_DutyLabel" )]

        dt3x = pd.merge(df_Final_sub, 
                        t3_wid, on= "name")\
            .set_index("name").T\
            .reset_index()\
            .assign(date_num = lambda df:[i[1] for i in df["index"]]) \
            .drop(columns ="index")\
            .set_index("date_num").T
        # dt3x

        # t3 = self.rp.dt3_mg.copy()
        # # t3
        # task_str = [r.DutyLabel if r.vals ==
        #             1 else "" for r in t3.itertuples()]
        # t3["task_str"] = task_str
        # t3_lim = t3.loc[:, ["name", "date_number",
        #                     "num_days"]].drop_duplicates()
        # duty_strs = []

        # for r in t3_lim.itertuples():
        #     r_name = r.name
        #     r_date_number = r.date_number
        #     r_num_days = r.num_days
        #     l_str = t3.query("name == @r_name and date_number == @r_date_number and num_days == @r_num_days")[
        #         "task_str"].to_list()
        #     duty_strs.append(self.comb_str(l_str))
        # t3_lim["duty_str"] = duty_strs
        # t3_lim = t3_lim.set_index(["date_number", "num_days"])
        # t3_wide = pd.DataFrame()

        # for name in tqdm(t3_lim["name"].unique(), disable=self.hideProgressBar):
        #     t3_wide[name] = t3_lim.query("name == @name")["duty_str"]

        # df_Final_sub = self.rp.x.df_FinalStatusAtLastMonth.copy(
        # ).loc[:, ["name", "LastDay_DutyLabel"]]
        # df_Final_sub.index = df_Final_sub["name"]
        # df_Final_sub = df_Final_sub.drop("name", axis=1)
        # df_Final_sub = df_Final_sub.fillna("")


        # idx = pd.MultiIndex.from_arrays(
        #     [[0], ["last_day"]], names=("date_number", "num_days"))
        # df_Final_sub.columns = idx

        # dt3x = pd.merge(df_Final_sub, t3_wide.T,
        #                 left_index=True, right_index=True)

        # for col in dt3x.columns:
        #     # print(col)
        #     dtx3_i = dt3x[col]
        #     dt3x[col] = [t if t is not np.nan else "" for t in dtx3_i]

        return dt3x

    def getComparisonTable(self):
        print("get! mada tukutteinai!")

    def getEachTaskTable(self):
        t3 = self.rp.dt3_mg.copy()
        t3_lim = t3.loc[:, ["date_number", "num_days", "tasks", "task_num", "count_in_task", "DutyLabel", "DutyClass",
                            "isNightWork", "task_ind", "dclass_ind"]].drop_duplicates()
        # t3_lim.head()
        name_list = []
        members_nums = []
        for r in t3_lim.itertuples():
            r_date_number = r.date_number
            r_dclass_ind = r.dclass_ind
            r_task_num = r.task_num
            r_count = r.count_in_task
            names = t3.query(
                "date_number == @r_date_number and dclass_ind == @r_dclass_ind and task_num == @r_task_num and count_in_task == @r_count and vals ==1")[
                "name"].to_list()
            members = len(names)
            name_list.append(self.comb_str(names))
            members_nums.append(members)
        t3_lim["names"] = name_list
        t3_lim["member_number"] = members_nums
        t3_sort = t3_lim.sort_values(
            ["date_number", "dclass_ind", "task_ind", "count_in_task"])

        return t3_sort

    def getMemberStats(self):
        # print("mada!")
        dt1 = self.rp.dt1_mg.query("vals==1")
        dt2 = self.rp.dt2_mg.query("vals==1")
        # fix
        # dt1.loc[:, "weekdays"] = dt1["num_days"].apply(
        dt1.loc[:, "weekdays"] = dt1["num_days"].copy().apply(
            lambda x: x.split("_")[1]).tolist()
        # fix
        # dt1.loc[:, "holiday_count"] = dt1["isHoliday"].apply(
        dt1.loc[:, "holiday_count"] = dt1["isHoliday"].copy().apply(
            lambda x: 1 if x else 0).tolist()
        sun_hol = [1 if r.weekdays ==
                   "Sunday" or r.isHoliday else 0 for r in dt1.itertuples()]
        dt1.loc[:, "sun_hol_count"] = sun_hol

        vals_wk = dt1.groupby(["name", "weekdays"]).sum()["vals"]
        vals_wk = vals_wk.unstack(fill_value=0)
        vals_wk = vals_wk.loc[:, [
            "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]]
        # vals_wk

        vals_total = dt1.groupby(["name"]).sum(
        ).loc[:, ["vals", "holiday_count", "sun_hol_count"]]
        vals_total.columns = ["total_count", "holiday_count", "sun_hol_count"]

        smr_dt1 = pd.merge(vals_total, vals_wk, on="name")

        dt2_workhour = dt2.groupby("name")\
            [["WorkHours"]].sum()
        dt2_n = dt2.groupby("name").sum()["isNightWork"]
        dt2_smr = pd.merge(smr_dt1, dt2_n, on="name")\
            .merge(dt2_workhour, on="name")

        dt2_sc = dt2\
            .query("vals==1")\
            .groupby(["name", "DutyClass"])["takability"]\
            .count()\
            .unstack()\
            .fillna(0)\
            .reset_index(drop=False)

        dt2_re = pd.merge(dt2_smr, dt2_sc, on="name", how="left")
        
        df0 = self.rp.x.df_MemberAbility.loc[:, ["name", "memo"]]
        # dt2_re2 = pd.merge(df0, dt2_re, on="name", how="left").fillna(0) 
        dt2_re2 = pd.merge(df0, dt2_re, on="name", how="left").fillna(0) \
            .assign(n_rest = lambda df: -df.MY -df.NN -df.Yy -df["ﾆB"] + self.rp.x.max_date)
            

        return dt2_re2

    def getTeamStats_byshift(self):
        # print("mada!")
        dt2 = self.rp.dt2_mg.query("vals == 1")
        dt2_smr = dt2.groupby(["date_number", "DutyClass", "group"]).sum()
        dt2_smr = dt2_smr["vals"].unstack(fill_value=0)

        d_dc2int = {
            "ﾆB":1,
            "NN":2,
            "MY":3,
            "Yy":4
        }

        df_dateclass = self.rp.x.df_weeks \
            .loc[:, ["date_number", "days4proc"]]\
            .copy()\
            .rename(columns = {"days4proc":"DateClass"})

        dt2_smr = dt2_smr\
            .copy()\
            .reset_index()\
            .assign(DutyClassInt = lambda df:[d_dc2int[k] for k in df.DutyClass])\
            .assign(ABTotalCount = lambda df: df.A + df["A/B"] + df.B) \
            .sort_values(["date_number", "DutyClassInt"]) \
            .merge(df_dateclass, on="date_number", how="left")\
            .drop(columns = "DutyClassInt")\
            .reindex(columns =["date_number", "DateClass", "DutyClass", "A", "A/B", "B", "ABTotalCount"])\
            .set_index(["date_number", "DateClass", "DutyClass"])

        return dt2_smr

    def getTeamStats_byduty(self):
        dt3 = self.rp.dt3_mg.query("vals ==1")
        # dt3.head().columns
        dt3_smr = dt3.groupby(
            ["date_number", "DutyLabel", "group"]).sum()["vals"]
        dt3_smr = dt3_smr.unstack(fill_value=0)
        
        df_dateclass = self.rp.x.df_weeks \
            .loc[:, ["date_number", "days4proc"]]\
            .copy()\
            .rename(columns = {"days4proc":"DateClass"})

        d_dl2int = {
            "ﾆB_L":1,
            "ﾆB_S":2,
            "ﾆB_I":3,
            "NN_S":4,
            "NN_I":5,
            "MY_S":6,
            "MY_I":7,
            "YY_S":8,
            "YY_I":9,
            "OC":100
        }

        dt3_smr = dt3_smr\
            .copy()\
            .reset_index()\
            .assign(DutyLabelInt = lambda df:[d_dl2int[k] for k in df.DutyLabel])\
            .assign(ABTotalCount = lambda df: df.A + df["A/B"] + df.B) \
            .sort_values(["date_number", "DutyLabelInt"]) \
            .merge(df_dateclass, on="date_number", how="left")\
            .drop(columns = "DutyLabelInt")\
            .reindex(columns =["date_number", "DateClass", "DutyLabel", "A", "A/B", "B", "ABTotalCount"])\
            .set_index(["date_number", "DateClass", "DutyLabel"])

        return dt3_smr

    def getFinalState(self):
        # print("unko2!")
        dt1 = self.rp.dt1_mg.copy() \
            .loc[:, ["date_number", "num_days", "days4proc", "name", "vals"]] \
            .drop_duplicates() \
            .set_index(["name", "date_number"])\
            .loc[:, "vals"]
        dt1_inv = dt1.unstack().apply(lambda x: (x - 1) ** 2).T

        # check finals rest
        max_date = max(dt1_inv.index)
        # max_date
        last_rest = []
        dt1_inv_c = dt1_inv.copy()
        for m in dt1_inv.columns:
            dt1_inv_c.loc[:, m] = dt1_inv.index * dt1_inv.loc[:, m]
            last_rest.append(int(dt1_inv_c[m].max()))

        final_stat = pd.DataFrame(columns=["last_rest"], index=dt1_inv.columns)
        final_stat["last_rest"] = last_rest
        final_stat["last_cont_work"] = max_date - final_stat["last_rest"]

        # check final duty

        dt3 = self.rp.dt3_mg.copy()
        dt3 = dt3.query(
            "date_number == @max_date and vals ==1").loc[:, ["name", "DutyLabel", "DutyClass"]]
        dt3_new = pd.DataFrame(columns=["name", "DutyLabel", "DutyClass"])
        for i, m in enumerate(dt3["name"].unique()):
            dl_l = dt3.query("name == @m")["DutyLabel"].to_list()
            dc_l = dt3.query("name == @m")["DutyClass"].to_list()
            row_m = [m, self.comb_str(l_str=dl_l), self.comb_str(dc_l)]
            dt3_new.loc[i, :] = row_m
        # dt3_new

        dt3 = dt3_new
        final_stat2 = pd.merge(
            final_stat, dt3, on="name", how="left").fillna("")
        # final_stat2

        # about night

        dt2 = self.rp.dt2_mg.copy().query("isNightWork == 1 and vals ==1") \
            .loc[:, ["date_number", "num_days", "days4proc", "name", "vals"]] \
            .drop_duplicates() \
            .set_index(["name", "date_number"]) \
            .loc[:, "vals"] \
            .unstack(fill_value=0)
        # dt2.head()
        dt2 = dt2.T.apply(lambda x: (x - 1) ** 2)
        # dt2.head()

        last_non_night = []
        for m in final_stat2["name"]:
            if m in dt2.columns:
                dt2[m] = dt2.index * dt2[m]
                last_non_night.append(int(dt2[m].max()))
            else:
                dt2[m] = max_date
        dt2_max = dt2 \
            .apply(max)\
            .reset_index()\
            .rename(columns={0: "last_non_night"})
        # dt2_max
        final_stat2 = pd.merge(final_stat2, dt2_max, on="name")
        final_stat2["last_cont_night"] = max_date - \
            final_stat2["last_non_night"]
        # return final_stat2

        final_stat3 = pd.merge(self.rp.x.df_FinalStatusAtLastMonth, final_stat2, on="name") \
                        .loc[:, ["name", "last_cont_work", "last_cont_night", "DutyLabel", "DutyClass"]]
        # #                     .copy()
        final_stat3.columns = ["name", "LastContinuousWorks", "LastContinuousNightWorks", "LastDay_DutyLabel",
                               "LastDay_DutyClass"]
        return final_stat3

    def get_c0_removed(self, df):
        df2 = df.copy()
        df2.columns = df2.columns.droplevel(0)
        return df2

    def get_c1_removed(self, df):
        df2 = df.copy()
        df2.columns = df2.columns.droplevel(1)
        return df2

    def add_tables2xlsx(self, f1, f2, rc_log_df=None):
        # print("unko!")

        shutil.copyfile(f1, f2)
        book = load_workbook(f2)

        book = self.replaceWs(book, "FinalStatusAtLastMonth",
                              df=self.finalstat, ind=False)

        book = self.replaceWs(book, "ShiftTable",
                              df=self.get_c0_removed(self.shifttable))
        # book = self.replaceWs(
        #     book, "DutyTable", df=self.get_c0_removed(self.dutytable))
        book = self.replaceWs(
            book, "DutyTable", df=self.dutytable)
        book = self.replaceWs(book, "EachTask", df=self.eachtask, ind=False)
        dfdt = self.datetable.copy()
        dfdt = self.get_c0_removed(dfdt)
        dfdt = self.get_c1_removed(dfdt)

        book = self.replaceWs(book, "DateTable", df=dfdt)
        book = self.replaceWs(book, "MemberStats", df=self.member_stat)

        if rc_log_df is not None:
            book = self.replaceWs(book, "ConditionLog", df=rc_log_df)
        # book = self.replaceWs(book, "TeamInfo_byDutyClass", df=self.teamstats_byshift)

        if "TeamInfo_byDutyClass" in book.sheetnames:
            del book["TeamInfo_byDutyClass"]
        if "TeamInfo_byDuty" in book.sheetnames:
            del book["TeamInfo_byDuty"]

        book.save(f2)

        with pd.ExcelWriter(f2, engine='openpyxl', mode="a") as writer:
            self.teamstats_byshift.to_excel(
                writer, sheet_name="TeamInfo_byDutyClass")
            self.teamstats_byduty.to_excel(
                writer, sheet_name="TeamInfo_byDuty")
            
        # change color
        wb = xl.load_workbook(f2)
        ws = wb["ShiftTable"]
        d_dc2col = {
            "NN":"fdb462",
            "ﾆB":"ffffb3",
            "MY":'bebada',
            "Yy":"80b1d3",
            "YY":"80b1d3",
            "OC":"d53e4f",
            "None":"f7f7f7"
        }

        d_dc2fill = dict()
        for k in d_dc2col.keys():
            print(k)
            d_dc2fill[k] = xl.styles.PatternFill(patternType="solid",
                                            fgColor=d_dc2col[k],
                                            bgColor=d_dc2col[k])
        for row in ws.iter_rows():
            # print(row[0].value)
            if row[0].value is not None:
                # print(row[0].value)
                for cell in row[1:]:
                    # print(cell.value)
                    if cell.value is None:
                        cell.fill = d_dc2fill["None"]
                    else:
                        # print(cell.value)
                        cell.fill = d_dc2fill[cell.value]
        
        ws2 = wb["DutyTable"]
        for row in ws2.iter_rows():
            # print(row[0].value)
            if row[0].value is not None:
                # print(row[0].value)
                for cell in row[1:]:
                    # print(cell.value)
                    if cell.value is None:
                        cell.fill = d_dc2fill["None"]
                    else:
                        val = cell.value[:2]
                        # print(cell.value)
                        cell.fill = d_dc2fill[val]

        wb.save(f2)

    def replaceWs(self, book, sheet, df, ind=True, col=True):
        if len(df.columns[0]) != 1:
            df.columns.labels = list(df.columns)

        if sheet in book.sheetnames:
            ws = book[sheet]

            # delete all 1st
            for r in ws:
                ws.delete_rows(1)

            # delete and write
            # 最終列がどういうわけかカウントされない。一行追加。
            ws.insert_rows(1)

            for indx, r in enumerate(dataframe_to_rows(df, index=ind, header=col)):
                # print(ind, r)
                ws.delete_rows(1)
                ws.append(r)
            return book

        else:
            book.create_sheet(sheet)
            ws = book[sheet]
            for r in dataframe_to_rows(df, index=ind, header=col):
                # print(r)
                ws.append(r)
            return book
